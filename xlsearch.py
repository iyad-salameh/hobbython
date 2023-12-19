import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import tkinter.font as tkFont
import os

def find_excel_files():
    """Finds all Excel files in the current directory."""
    return [f for f in os.listdir('.') if f.endswith('.xlsx')]

def load_workbooks():
    """Loads all workbooks and stores them in a dictionary."""
    workbooks = {}
    for file in find_excel_files():
        try:
            workbooks[file] = openpyxl.load_workbook(file)
        except Exception as e:
            print(f"Error loading {file}: {e}")
    return workbooks

def search_excel(keyword, workbooks):
    """Searches for the keyword in all loaded workbooks."""
    results = []
    for file, workbook in workbooks.items():
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                if keyword.lower() in [str(cell).lower() for cell in row]:
                    results.append((file, sheet_name, row))
    return results

def adjust_column_widths(tree):
    """Adjusts the column widths based on the data."""
    tree.update()  # Update the treeview to ensure all data is loaded
    for col in tree['columns']:
        # Find the maximum width of data in each column
        col_width = font.measure(col)  # Start with the width of the column header
        for row in tree.get_children():
            cell_width = font.measure(tree.set(row, col))
            col_width = max(col_width, cell_width)
        tree.column(col, width=col_width)

def on_search():
    """Handles the search event."""
    keyword = search_entry.get().strip()
    if not keyword:
        messagebox.showinfo("Search", "Please enter a keyword to search.")
        return

    results = search_excel(keyword, loaded_workbooks)
    tree.delete(*tree.get_children())
    for file, sheet, row in results:
        tree.insert('', tk.END, values=[file, sheet] + list(row))
    adjust_column_widths(tree)  # Adjust columns after the search

loaded_workbooks = load_workbooks()  # Load workbooks at start

# GUI Setup
root = tk.Tk()
root.title("Excel Sheet Searcher")

font = tkFont.Font(family="TkDefaultFont", size=10) #adjust size thing

search_entry = ttk.Entry(root)
search_entry.pack(pady=10)

search_button = ttk.Button(root, text="Search", command=on_search)
search_button.pack(pady=10)

# Create a Treeview widget with a horizontal scrollbar
tree = ttk.Treeview(root, columns=("File", "Sheet") + tuple(f"Column {i+1}" for i in range(10)), show='headings')
scrollbar = ttk.Scrollbar(root, orient="horizontal", command=tree.xview)
tree.configure(xscrollcommand=scrollbar.set)
scrollbar.pack(side="bottom", fill="x")

# Set the headings
for i, col_name in enumerate(["File", "Sheet"] + [f"Column {i+1}" for i in range(10)]):
    tree.heading(col_name, text=col_name)
    tree.column(col_name, width=100)  # Initial width, will be adjusted

tree.pack(expand=True, fill='both')

# Run the GUI
root.mainloop()

