import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import tkinter.font as tkFont
import os
import time


def get_column_letter(col_num):
    """Converts a column number (e.g., 1, 2, 3, ...) into a column letter (e.g., A, B, C, ...)."""
    string = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        string = chr(65 + remainder) + string
    return string

def find_excel_files():
    """Finds all Excel files in the current directory."""
    return [f for f in os.listdir('.') if f.lower().endswith('.xlsx')]

def load_workbooks():
    """Loads all workbooks and stores them in a dictionary."""
    workbooks = {}
    for file in find_excel_files():
        try:
            workbooks[file] = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"Error loading {file}: {e}")
    return workbooks

def search_excel(keyword, workbooks):
    """Searches for the keyword in all loaded workbooks."""
    keyword_lower = keyword.lower()
    results = []
    for file, workbook in workbooks.items():
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_lower = [str(cell).lower() if cell is not None else "" for cell in row]
                for cell in row_lower:
                    if keyword_lower in cell:
                        results.append((file, sheet_name, row[:MAX_COLUMN_DISPLAY]))
    return results

#adjust the column widths according to cell entry or something..
def adjust_column_widths(tree, font):
    """Adjusts the column widths based on the data."""
    for col in tree["columns"]:
        max_width = font.measure(col)
        for row in tree.get_children():
            cell_value = tree.set(row, col)
            max_width = max(max_width, font.measure(cell_value))
        tree.column(col, width=max_width + 10)  # Add a little extra space

# GUI Setup
root = tk.Tk()
root.title("Excel Sheet Searcher")
root.geometry("800x600")  # Adjust the size of the window as needed

# Define the font for measuring text width
font = tkFont.Font(family="TkDefaultFont", size=10)  # Adjust size as needed

search_entry = ttk.Entry(root)
search_entry.pack(pady=10)

search_button = ttk.Button(root, text="Search", command=lambda: on_search())
search_button.pack(pady=10)

# Create a Treeview widget with a horizontal scrollbar
MAX_COLUMN_DISPLAY = 10
columns = ("File", "Sheet") + tuple(get_column_letter(i+1) for i in range(MAX_COLUMN_DISPLAY))
tree = ttk.Treeview(root, columns=columns, show='headings')
scrollbar = ttk.Scrollbar(root, orient="horizontal", command=tree.xview)
tree.configure(xscrollcommand=scrollbar.set)
scrollbar.pack(side="bottom", fill="x")

for col_name in columns:
    tree.heading(col_name, text=col_name)
    tree.column(col_name, width=tkFont.Font().measure(col_name))  # Initial width, will be adjusted

tree.pack(expand=True, fill='both')

loaded_workbooks = load_workbooks()  # Load workbooks at start

def on_search():
    """Handles the search event."""
    keyword = search_entry.get().strip()
    if not keyword:
        messagebox.showinfo("Search", "Please enter a keyword to search.")
        return
    try:
        results = search_excel(keyword, loaded_workbooks)
        tree.delete(*tree.get_children())
        for file, sheet, row in results:
            tree.insert('', tk.END, values=(file, sheet) + tuple(row))
        adjust_column_widths(tree, font)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred during search: {e}")

# Function to copy the selected row to the clipboard
def copy_to_clipboard(event):
    selected_item = tree.focus()  # Get the selected item
    if selected_item:
        row_data = tree.item(selected_item, 'values')
        row_text = '\t'.join(map(str, row_data))  # Convert tuple to tab-separated string
        root.clipboard_clear()
        root.clipboard_append(row_text)
        
# Function to show file properties in a popup window
def show_properties():
    selected_item = tree.focus()  # Get the selected item
    if selected_item:
        file_name = tree.item(selected_item, 'values')[0]  # Assuming first value is the file name
        properties = get_file_properties(file_name)

        # Create a popup window
        popup = tk.Toplevel()
        popup.title("File Properties")
        for key, value in properties.items():
            ttk.Label(popup, text=f"{key}: {value}").pack()
        
#function to get file properties
def get_file_properties(filepath):
    properties = {
        'Path': filepath,
        'Size': f"{os.path.getsize(filepath)} bytes",
        'Created': time.ctime(os.path.getctime(filepath)),
        'Modified': time.ctime(os.path.getmtime(filepath))
    }
    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        properties['Author'] = workbook.properties.author or "Unknown"
    except Exception:
        properties['Author'] = "Unknown"
    return properties

# Create a context menu for copying and properties
context_menu = tk.Menu(root, tearoff=0)
context_menu.add_command(label="Properties", command=show_properties)
context_menu.add_command(label="Copy to Clipboard", command=lambda: copy_to_clipboard(None))

# Function to display the context menu
def on_right_click(event):
    try:
        context_menu.tk_popup(event.x_root, event.y_root)
    finally:
        context_menu.grab_release()

# Bind the right-click event
tree.bind("<Button-3>", on_right_click)

# Bind Enter key to search
search_entry.bind('<Return>', lambda event: on_search())
def on_search(event=None):
    keyword = search_entry.get().strip()
    if not keyword:
        messagebox.showinfo("Search", "Please enter a keyword to search.")
        return
    try:
        results = search_excel(keyword, loaded_workbooks)
        tree.delete(*tree.get_children())
        for file, sheet, row in results:
            tree.insert('', tk.END, values=(file, sheet) + tuple(row))
        adjust_column_widths(tree, font)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred during search: {e}")

root.mainloop()
